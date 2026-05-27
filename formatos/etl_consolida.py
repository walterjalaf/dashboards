#!/usr/bin/env python3
"""
ETL Don Américo — Consolida los archivos mensuales en un único CSV
apto para alimentar la Google Sheet 'Resumen_Mensual_DonAmerico_2026'.

USO:
    python etl_consolida.py /carpeta/con/excels/  /salida/dataset.csv

Detecta automáticamente:
  - Formato legacy (sheet ESTADOS, ENE/FEB)
  - Formato nuevo (sheet ECONOMICO, MAR/ABR en adelante)
  - Rubros nuevos que aparezcan (se acumulan en rubro_otros)

Reglas:
  - KPIs principales se leen de DATA STUDIO (capa estable)
  - Rubros se leen por rótulo (col A) en ESTADOS/ECONOMICO
  - total_egresos se calcula, NO se lee
  - resultado_economico se lee y se valida contra ventas - total_egresos
"""
import sys, csv, json, re, warnings
from pathlib import Path
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

MES_MAP = {
    'ENERO':1, 'FEBRERO':2, 'MARZO':3, 'ABRIL':4, 'MAYO':5, 'JUNIO':6,
    'JULIO':7, 'AGOSTO':8, 'SEPTIEMBRE':9, 'OCTUBRE':10,
    'NOVIEMBRE':11, 'DICIEMBRE':12,
}
RUBROS_PRINCIPALES = ['CARNES','FIAMBRES','COMESTIBLES','BEBIDAS','VERDURAS',
                      'PANIFICACION','POLLO','LACTEOS','LIMPIEZA','KIOSCO']
RUBROS_OTROS_CONOCIDOS = ['CERDO','ROTISERIA','PERFUMERIA','FRESCOS',
                          'CONGELADOS','VARIOS','DESCARTABLES']

# Mapeo etiqueta DATA STUDIO → columna del CSV
DS_MAP = {
    'VENTAS': 'ventas',
    'COSTO DE VENTAS': 'costo_ventas',
    'GASTOS': 'gastos_empresa',
    'IMPUESTOS': 'impuestos',
    'SERVICIOS': 'servicios',
    'EMPLEADOS': 'sueldos',
    'GASTOS Y COMISIONES BANCARIAS': 'gastos_bancarios',
    'GASTOS Y COMISIONES MERCADO PAGO': 'gastos_mercado_pago',
    'GASTOS Y COMISIONES POSNET': 'gastos_posnet',
    'EXTRAORDINARIOS': 'extraordinarios',
    'COMPRAS': 'compras',
    'STOCK AL CIERRE': 'stock_cierre',
    'SALDO PROVEEDORES': 'saldo_proveedores',
    'SALDO CLIENTES': 'saldo_clientes',
    'RESULTADO ECONOMICO': 'resultado_economico',
}


def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    return 0.0


def detectar_mes_año(filename):
    """Extrae mes y año del nombre del archivo."""
    name = filename.upper()
    mes = None
    for m in MES_MAP:
        if m in name:
            mes = m
            break
    año_match = re.search(r'(20\d{2})', name)
    año = int(año_match.group(1)) if año_match else None
    return mes, año


def hoja_economica(wb):
    """Devuelve la hoja con el desglose de rubros (ESTADOS o ECONOMICO)."""
    for cand in ('ECONOMICO', 'ESTADOS'):
        if cand in wb.sheetnames:
            return wb[cand]
    raise ValueError("No se encontró ESTADOS ni ECONOMICO")


def leer_data_studio(ws):
    """Lee DATA STUDIO (A:B) y devuelve dict etiqueta→valor."""
    out = {}
    for r in range(1, 25):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k and isinstance(k, str):
            out[k.strip().upper()] = num(v)
    return out


def leer_rubros(ws):
    """Lee rubros por rótulo en col A.
    Comienza después de 'CLASIF DE VENTAS POR RUBROS' y termina en
    'CLASIF DE VENTAS POR FORMAS DE COBROS' o 'TOTAL EGRESOS'.
    Eso evita confundir formas de cobro (PESOS, TARJETAS, TRANSF MP, CC) con rubros.
    """
    out = {}
    en_seccion_rubros = False
    for r in range(1, 60):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if not (a and isinstance(a, str)):
            continue
        key = a.strip().upper()
        if key == 'CLASIF DE VENTAS POR RUBROS':
            en_seccion_rubros = True
            continue
        if key in ('CLASIF DE VENTAS POR FORMAS DE COBROS',
                   'TOTAL EGRESOS', 'CLASIF DE EGRESOS POR RUBROS'):
            break
        if en_seccion_rubros and isinstance(b, (int, float)):
            out[key] = num(b)
    return out


def procesar_archivo(filepath):
    """Devuelve dict con una fila normalizada del mes."""
    mes_nom, año = detectar_mes_año(Path(filepath).name)
    if not mes_nom or not año:
        raise ValueError(f"No se pudo extraer mes/año de {filepath}")

    wb = load_workbook(filepath, data_only=True)
    ds_vals = leer_data_studio(wb['DATA STUDIO'])
    rubros = leer_rubros(hoja_economica(wb))
    wb.close()

    fila = {
        'mes_nombre': mes_nom,
        'mes_num': MES_MAP[mes_nom],
        'año': año,
    }
    # KPIs principales
    for ds_key, col in DS_MAP.items():
        fila[col] = ds_vals.get(ds_key, 0)

    # Total egresos calculado
    fila['total_egresos'] = sum(fila[k] for k in
        ('costo_ventas','gastos_empresa','impuestos','servicios','sueldos',
         'gastos_bancarios','gastos_mercado_pago','gastos_posnet'))
    fila['margen_neto'] = (fila['resultado_economico'] / fila['ventas'] * 100
                           if fila['ventas'] else 0)

    # Rubros principales
    for k in RUBROS_PRINCIPALES:
        fila[f'rubro_{k.lower()}'] = rubros.get(k, 0)

    # Otros: agrupado + desglose
    otros_total = 0
    rubros_otros_detectados = set(RUBROS_OTROS_CONOCIDOS)
    # Detectar rubros nuevos que no estén en los principales ni en otros conocidos
    for k in rubros:
        if k not in RUBROS_PRINCIPALES and k not in RUBROS_OTROS_CONOCIDOS:
            rubros_otros_detectados.add(k)
            print(f"  AVISO: rubro nuevo detectado '{k}' = {rubros[k]:,.2f} — agregado a 'otros'")

    for k in rubros_otros_detectados:
        v = rubros.get(k, 0)
        otros_total += v
        fila[f'rubro_otros_{k.lower()}'] = v
    fila['rubro_otros'] = otros_total

    # Validación cruzada
    diff = abs(fila['ventas'] - fila['total_egresos'] - fila['resultado_economico'])
    if diff > 1:
        print(f"  ⚠️  {mes_nom}: descuadre de {diff:,.2f} entre ventas-egresos vs resultado")

    return fila


def main():
    if len(sys.argv) < 3:
        print("Uso: python etl_consolida.py <carpeta_excels> <salida.csv>")
        sys.exit(1)

    carpeta = Path(sys.argv[1])
    salida = sys.argv[2]
    archivos = sorted(carpeta.glob('RESUMEN_MENSUAL_*.xlsx'))
    if not archivos:
        print(f"No se encontraron archivos RESUMEN_MENSUAL_*.xlsx en {carpeta}")
        sys.exit(1)

    print(f"Procesando {len(archivos)} archivo(s):")
    filas = []
    for f in archivos:
        print(f"  → {f.name}")
        try:
            filas.append(procesar_archivo(str(f)))
        except Exception as e:
            print(f"     ERROR: {e}")

    if not filas:
        print("Sin filas procesadas. Abortando.")
        sys.exit(1)

    # Ordenar por año, mes
    filas.sort(key=lambda r: (r['año'], r['mes_num']))

    # Unir todas las columnas (algunas filas pueden tener rubros distintos)
    cols = ['mes_nombre','mes_num','año']
    cols += ['ventas','costo_ventas','gastos_empresa','impuestos','servicios',
             'sueldos','gastos_bancarios','gastos_mercado_pago','gastos_posnet',
             'extraordinarios','compras','stock_cierre','saldo_proveedores',
             'saldo_clientes','resultado_economico','total_egresos','margen_neto']
    cols += [f'rubro_{k.lower()}' for k in RUBROS_PRINCIPALES]
    cols += ['rubro_otros']
    # detectar todas las columnas rubro_otros_* presentes en cualquier fila
    otros_cols = set()
    for fila in filas:
        for k in fila:
            if k.startswith('rubro_otros_'):
                otros_cols.add(k)
    cols += sorted(otros_cols)

    with open(salida, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for fila in filas:
            # rellenar columnas faltantes con 0
            for c in cols:
                fila.setdefault(c, 0)
            w.writerow(fila)

    print(f"\n✓ CSV consolidado: {salida}")
    print(f"  {len(filas)} fila(s), {len(cols)} columna(s)")
    print(f"\nResumen:")
    print(f"  {'Mes':10}{'Ventas':>18}{'Egresos':>18}{'Resultado':>15}")
    for r in filas:
        print(f"  {r['mes_nombre']:10}{r['ventas']:>18,.2f}{r['total_egresos']:>18,.2f}{r['resultado_economico']:>15,.2f}")


if __name__ == '__main__':
    main()
